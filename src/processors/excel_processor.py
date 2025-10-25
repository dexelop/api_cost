"""
Excel 파일 프로세서

Excel 파일의 데이터를 추출하여 텍스트로 변환합니다.
"""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .base import BaseFileProcessor, ProcessedFile


class ExcelFileProcessor(BaseFileProcessor):
    """
    Excel 파일 프로세서

    openpyxl과 pandas를 사용하여 Excel 파일의 데이터를 추출합니다.
    """

    def __init__(self):
        """Excel 프로세서 초기화"""
        super().__init__()
        self.supported_extensions = [".xlsx", ".xlsm", ".xls", ".csv"]

    def extract_sheet_data(self, file_path: Path, sheet_name: str = None) -> str:
        """
        Excel 시트의 데이터를 텍스트로 추출

        Args:
            file_path: Excel 파일 경로
            sheet_name: 추출할 시트 이름 (None이면 모든 시트)

        Returns:
            str: 추출된 데이터를 텍스트로 변환한 문자열
        """
        all_text = []

        # CSV 파일은 pandas로 직접 읽기
        if file_path.suffix.lower() == ".csv":
            df = pd.read_csv(file_path)
            return df.to_string(index=False)

        # Excel 파일 처리
        try:
            # 모든 시트 읽기
            excel_file = pd.ExcelFile(file_path)

            for sheet in excel_file.sheet_names:
                # 특정 시트만 읽기 (지정된 경우)
                if sheet_name and sheet != sheet_name:
                    continue

                # 시트 데이터 읽기
                df = pd.read_excel(file_path, sheet_name=sheet)

                # 시트 제목 추가
                all_text.append(f"=== Sheet: {sheet} ===\n")

                # DataFrame을 문자열로 변환
                all_text.append(df.to_string(index=False))
                all_text.append("\n\n")

        except Exception as e:
            all_text.append(f"[시트 데이터 추출 실패: {str(e)}]")

        return "\n".join(all_text)

    def get_excel_metadata(self, file_path: Path) -> dict:
        """
        Excel 파일의 메타데이터 추출

        Args:
            file_path: Excel 파일 경로

        Returns:
            dict: Excel 메타데이터
        """
        # 기본 파일 메타데이터
        metadata = self.get_file_metadata(file_path)

        # CSV 파일은 단순 처리
        if file_path.suffix.lower() == ".csv":
            try:
                df = pd.read_csv(file_path)
                metadata.update(
                    {
                        "sheet_count": 1,
                        "row_count": len(df),
                        "column_count": len(df.columns),
                        "total_cells": len(df) * len(df.columns),
                    }
                )
            except Exception:
                pass
            return metadata

        # Excel 파일 메타데이터 추출
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)

            # 시트 정보
            sheet_names = workbook.sheetnames
            total_rows = 0
            total_cols = 0

            # 각 시트의 행/열 수 계산
            for sheet_name in sheet_names:
                sheet = workbook[sheet_name]
                total_rows += sheet.max_row
                total_cols = max(total_cols, sheet.max_column)

            metadata.update(
                {
                    "sheet_count": len(sheet_names),
                    "sheet_names": sheet_names,
                    "total_rows": total_rows,
                    "max_columns": total_cols,
                    "total_cells": total_rows * total_cols,
                }
            )

            workbook.close()

        except Exception as e:
            metadata["extraction_error"] = str(e)

        return metadata

    def process(self, file_path: Path) -> ProcessedFile:
        """
        Excel 파일을 처리하여 데이터 추출

        Args:
            file_path: 처리할 Excel 파일 경로

        Returns:
            ProcessedFile: 처리된 파일 정보
        """
        try:
            # 파일 검증
            self.validate_file(file_path)

            # 데이터 추출
            content = self.extract_sheet_data(file_path)

            # 메타데이터 추출
            metadata = self.get_excel_metadata(file_path)

            # 텍스트 관련 메타데이터 추가
            metadata.update(
                {
                    "character_count": len(content),
                    "word_count": len(content.split()),
                }
            )

            return ProcessedFile(
                file_path=file_path,
                file_type="excel",
                content=content,
                metadata=metadata,
                error=None,
            )

        except Exception as e:
            # 에러 발생 시 에러 정보와 함께 반환
            return ProcessedFile(
                file_path=file_path,
                file_type="excel",
                content="",
                metadata=self.get_file_metadata(file_path)
                if file_path.exists()
                else {},
                error=str(e),
            )
