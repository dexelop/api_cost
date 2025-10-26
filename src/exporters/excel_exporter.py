"""
Excel 내보내기 모듈

비용 추정 결과를 Excel 파일로 내보냅니다.
"""

import io
from typing import List
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.processors.base import ProcessedFile
from src.pricing.calculator import CostEstimate
from src.tokenizers.file_tokenizer import FileTokenizer


class ExcelExporter:
    """
    Excel 내보내기 클래스

    비용 추정 결과를 Excel 형식으로 변환합니다.
    """

    def __init__(self):
        """Excel 내보내기 초기화"""
        self.tokenizer = FileTokenizer()

        # 스타일 정의
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def export_workbook(
        self,
        processed_files: List[ProcessedFile],
        estimates: List[CostEstimate],
        output_ratio: float,
    ) -> bytes:
        """
        전체 워크북 생성

        Args:
            processed_files: 처리된 파일 리스트
            estimates: 비용 추정 결과 리스트
            output_ratio: 출력 토큰 비율

        Returns:
            bytes: Excel 파일 바이트
        """
        wb = Workbook()

        # 기본 시트 제거
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # 요약 시트
        self._create_summary_sheet(wb, processed_files, estimates, output_ratio)

        # 파일 정보 시트
        self._create_files_sheet(wb, processed_files)

        # 비용 비교 시트
        self._create_costs_sheet(wb, estimates, output_ratio)

        # BytesIO로 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    def _create_summary_sheet(
        self,
        wb: Workbook,
        processed_files: List[ProcessedFile],
        estimates: List[CostEstimate],
        output_ratio: float,
    ):
        """
        요약 시트 생성

        Args:
            wb: 워크북
            processed_files: 처리된 파일 리스트
            estimates: 비용 추정 결과 리스트
            output_ratio: 출력 토큰 비율
        """
        ws = wb.create_sheet("요약", 0)

        # 타이틀
        ws["A1"] = "LLM API Cost Calculator"
        ws["A1"].font = Font(size=16, bold=True, color="4472C4")
        ws.merge_cells("A1:D1")

        # 생성 시간
        ws["A2"] = f"생성 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells("A2:D2")

        # 토큰 정보
        total_tokens = sum(
            self.tokenizer.count_tokens_from_processed_file(pf).token_count
            for pf in processed_files
        )
        estimated_output = int(total_tokens * output_ratio)

        ws["A4"] = "토큰 분석"
        ws["A4"].font = Font(bold=True, size=12)

        ws["A5"] = "총 입력 토큰"
        ws["B5"] = total_tokens
        ws["A6"] = "예상 출력 토큰"
        ws["B6"] = estimated_output
        ws["A7"] = "출력 비율"
        ws["B7"] = f"{output_ratio:.1%}"
        ws["A8"] = "총 토큰"
        ws["B8"] = total_tokens + estimated_output

        # 비용 정보
        ws["A10"] = "가장 저렴한 모델"
        ws["A10"].font = Font(bold=True, size=12)

        if estimates:
            cheapest = estimates[0]
            ws["A11"] = "모델"
            ws["B11"] = cheapest.model.model_name
            ws["A12"] = "제공업체"
            ws["B12"] = cheapest.model.provider.upper()
            ws["A13"] = "총 비용"
            ws["B13"] = cheapest.total_cost
            ws["B13"].number_format = "$#,##0.000000"

        # 파일 정보
        ws["A15"] = "업로드된 파일"
        ws["A15"].font = Font(bold=True, size=12)
        ws["A16"] = f"{len(processed_files)}개 파일"

        # 컬럼 너비 조정
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    def _create_files_sheet(self, wb: Workbook, processed_files: List[ProcessedFile]):
        """
        파일 정보 시트 생성

        Args:
            wb: 워크북
            processed_files: 처리된 파일 리스트
        """
        ws = wb.create_sheet("파일 정보")

        # 헤더
        headers = ["파일명", "파일 타입", "파일 크기 (MB)", "토큰 수", "문자 수"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border

        # 데이터
        for row, pfile in enumerate(processed_files, 2):
            token_count = self.tokenizer.count_tokens_from_processed_file(pfile)

            ws.cell(row, 1, pfile.metadata.get("file_name", "Unknown"))
            ws.cell(row, 2, pfile.file_type)
            ws.cell(row, 3, pfile.metadata.get("file_size_mb", 0))
            ws.cell(row, 4, token_count.token_count)
            ws.cell(row, 5, len(pfile.content))

            # 테두리 적용
            for col in range(1, 6):
                ws.cell(row, col).border = self.border

        # 컬럼 너비 조정
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15

    def _create_costs_sheet(
        self, wb: Workbook, estimates: List[CostEstimate], output_ratio: float
    ):
        """
        비용 비교 시트 생성

        Args:
            wb: 워크북
            estimates: 비용 추정 결과 리스트
            output_ratio: 출력 토큰 비율
        """
        ws = wb.create_sheet("비용 비교")

        # 헤더
        headers = [
            "제공업체",
            "모델명",
            "모델 ID",
            "입력 토큰",
            "출력 토큰",
            "입력 가격 ($/1K)",
            "출력 가격 ($/1K)",
            "입력 비용 ($)",
            "출력 비용 ($)",
            "총 비용 ($)",
            "Context 윈도우",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border

        # 데이터
        for row, est in enumerate(estimates, 2):
            ws.cell(row, 1, est.model.provider.upper())
            ws.cell(row, 2, est.model.model_name)
            ws.cell(row, 3, est.model.model_id)
            ws.cell(row, 4, est.input_tokens)
            ws.cell(row, 5, est.output_tokens)
            ws.cell(row, 6, est.model.input_price)
            ws.cell(row, 7, est.model.output_price)
            ws.cell(row, 8, est.input_cost)
            ws.cell(row, 9, est.output_cost)
            ws.cell(row, 10, est.total_cost)
            ws.cell(row, 11, est.model.context_window)

            # 숫자 포맷
            ws.cell(row, 6).number_format = "$#,##0.0000"
            ws.cell(row, 7).number_format = "$#,##0.0000"
            ws.cell(row, 8).number_format = "$#,##0.000000"
            ws.cell(row, 9).number_format = "$#,##0.000000"
            ws.cell(row, 10).number_format = "$#,##0.000000"
            ws.cell(row, 11).number_format = "#,##0"

            # 테두리 적용
            for col in range(1, 12):
                ws.cell(row, col).border = self.border

        # 컬럼 너비 조정
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 15
        ws.column_dimensions["J"].width = 15
        ws.column_dimensions["K"].width = 15
