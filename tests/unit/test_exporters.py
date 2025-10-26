"""
내보내기 모듈에 대한 유닛 테스트

CSV 내보내기 기능을 검증합니다.
"""

from pathlib import Path

import pytest
import pandas as pd

from src.exporters.csv_exporter import CSVExporter
from src.exporters.excel_exporter import ExcelExporter
from src.exporters.json_exporter import JSONExporter
from src.processors.base import ProcessedFile
from src.pricing.calculator import PriceCalculator


class TestCSVExporter:
    """CSVExporter에 대한 테스트"""

    def test_export_file_tokens(self, tmp_path):
        """파일 토큰 정보 CSV 내보내기 테스트"""
        exporter = CSVExporter()

        # 테스트용 ProcessedFile 생성
        test_file = ProcessedFile(
            file_path=tmp_path / "test.txt",
            file_type="text",
            content="This is test content for token counting.",
            metadata={
                "file_name": "test.txt",
                "file_size_mb": 0.001,
            },
        )

        # CSV 내보내기
        csv_output = exporter.export_file_tokens([test_file])

        # CSV가 문자열이어야 함
        assert isinstance(csv_output, str)
        assert len(csv_output) > 0

        # CSV 파싱 가능 확인
        import io

        df = pd.read_csv(io.StringIO(csv_output))
        assert len(df) == 1
        assert "파일명" in df.columns
        assert "파일 타입" in df.columns
        assert "토큰 수" in df.columns

        # 값 확인
        assert df["파일명"].iloc[0] == "test.txt"
        assert df["파일 타입"].iloc[0] == "text"
        assert df["토큰 수"].iloc[0] > 0

    def test_export_cost_estimates(self):
        """비용 추정 CSV 내보내기 테스트"""
        exporter = CSVExporter()
        calculator = PriceCalculator()

        # 비용 계산
        estimate = calculator.calculate_cost("gpt-4", input_tokens=1000, output_tokens=500)

        # CSV 내보내기
        csv_output = exporter.export_cost_estimates([estimate], output_ratio=0.5)

        # CSV가 문자열이어야 함
        assert isinstance(csv_output, str)
        assert len(csv_output) > 0

        # CSV 파싱 가능 확인
        import io

        df = pd.read_csv(io.StringIO(csv_output))
        assert len(df) == 1
        assert "제공업체" in df.columns
        assert "모델명" in df.columns
        assert "총 비용 ($)" in df.columns
        assert "입력 토큰" in df.columns
        assert "출력 토큰" in df.columns

        # 값 확인
        assert df["제공업체"].iloc[0] == "OPENAI"
        assert df["입력 토큰"].iloc[0] == 1000
        assert df["출력 토큰"].iloc[0] == 500

    def test_export_combined(self, tmp_path):
        """통합 CSV 내보내기 테스트"""
        exporter = CSVExporter()
        calculator = PriceCalculator()

        # 테스트용 파일
        test_file = ProcessedFile(
            file_path=tmp_path / "test.txt",
            file_type="text",
            content="Test content",
            metadata={"file_name": "test.txt", "file_size_mb": 0.001},
        )

        # 비용 계산
        estimate = calculator.calculate_cost("gpt-4o-mini", input_tokens=1000)

        # 통합 CSV 내보내기
        csv_output = exporter.export_combined([test_file], [estimate], output_ratio=0.3)

        # CSV가 문자열이어야 함
        assert isinstance(csv_output, str)
        assert len(csv_output) > 0

        # CSV 파싱 가능 확인
        import io

        df = pd.read_csv(io.StringIO(csv_output))

        # 여러 섹션이 포함되어 있어야 함
        assert "구분" in df.columns
        assert "항목" in df.columns
        assert "값" in df.columns

        # 파일 섹션 확인
        file_rows = df[df["구분"] == "파일"]
        assert len(file_rows) > 0

        # 비용 섹션 확인
        cost_rows = df[df["구분"] == "비용"]
        assert len(cost_rows) > 0

        # 메타 섹션 확인
        meta_rows = df[df["구분"] == "메타"]
        assert len(meta_rows) > 0

    def test_export_multiple_files(self, tmp_path):
        """여러 파일 내보내기 테스트"""
        exporter = CSVExporter()

        # 여러 테스트 파일 생성
        files = [
            ProcessedFile(
                file_path=tmp_path / "file1.txt",
                file_type="text",
                content="File 1",
                metadata={"file_name": "file1.txt", "file_size_mb": 0.001},
            ),
            ProcessedFile(
                file_path=tmp_path / "file2.md",
                file_type="text",
                content="File 2",
                metadata={"file_name": "file2.md", "file_size_mb": 0.002},
            ),
        ]

        # CSV 내보내기
        csv_output = exporter.export_file_tokens(files)

        # CSV 파싱
        import io

        df = pd.read_csv(io.StringIO(csv_output))

        # 2개 파일 모두 포함되어야 함
        assert len(df) == 2
        assert "file1.txt" in df["파일명"].values
        assert "file2.md" in df["파일명"].values

    def test_export_multiple_models(self):
        """여러 모델 비교 내보내기 테스트"""
        exporter = CSVExporter()
        calculator = PriceCalculator()

        # 여러 모델 비용 비교
        models = ["gpt-4o-mini", "claude-3-haiku", "gemini-1.5-flash"]
        estimates = calculator.compare_models(models, input_tokens=1000, output_tokens=500)

        # CSV 내보내기
        csv_output = exporter.export_cost_estimates(estimates, output_ratio=0.5)

        # CSV 파싱
        import io

        df = pd.read_csv(io.StringIO(csv_output))

        # 3개 모델 모두 포함되어야 함
        assert len(df) == 3

        # 비용 순 정렬 확인
        # CSV에서 읽어온 값은 이미 float일 수도 있고 문자열일 수도 있음
        costs = []
        for val in df["총 비용 ($)"].values:
            if isinstance(val, str):
                costs.append(float(val.replace("$", "")))
            else:
                costs.append(float(val))
        assert all(costs[i] <= costs[i + 1] for i in range(len(costs) - 1))

    def test_to_bytes(self):
        """CSV 바이트 변환 테스트"""
        exporter = CSVExporter()

        csv_string = "Column1,Column2\nValue1,Value2"
        csv_bytes = exporter.to_bytes(csv_string)

        # bytes 타입이어야 함
        assert isinstance(csv_bytes, bytes)

        # UTF-8 BOM이 포함되어야 함 (Excel 한글 지원)
        assert csv_bytes.startswith(b"\xef\xbb\xbf")

        # 디코딩 가능 확인
        decoded = csv_bytes.decode("utf-8-sig")
        assert decoded == csv_string

    def test_export_with_korean(self, tmp_path):
        """한글 파일명 내보내기 테스트"""
        exporter = CSVExporter()

        # 한글 파일명
        test_file = ProcessedFile(
            file_path=tmp_path / "테스트.txt",
            file_type="text",
            content="한글 테스트 내용",
            metadata={"file_name": "테스트.txt", "file_size_mb": 0.001},
        )

        # CSV 내보내기
        csv_output = exporter.export_file_tokens([test_file])

        # UTF-8 인코딩 확인
        import io

        df = pd.read_csv(io.StringIO(csv_output))
        assert df["파일명"].iloc[0] == "테스트.txt"

        # bytes 변환 후 디코딩 가능 확인
        csv_bytes = exporter.to_bytes(csv_output)
        decoded = csv_bytes.decode("utf-8-sig")
        assert "테스트.txt" in decoded


class TestExcelExporter:
    """ExcelExporter에 대한 테스트"""

    def test_export_workbook(self, tmp_path):
        """Excel 워크북 내보내기 테스트"""
        exporter = ExcelExporter()
        calculator = PriceCalculator()

        # 테스트 파일
        test_file = ProcessedFile(
            file_path=tmp_path / "test.txt",
            file_type="text",
            content="Test content for Excel export",
            metadata={"file_name": "test.txt", "file_size_mb": 0.001},
        )

        # 비용 계산
        estimates = calculator.compare_models(
            ["gpt-4o-mini", "claude-3-haiku"], input_tokens=1000, output_tokens=500
        )

        # Excel 내보내기
        excel_bytes = exporter.export_workbook([test_file], estimates, output_ratio=0.5)

        # bytes 타입 확인
        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 0

        # Excel 파일 구조 확인 (openpyxl로 읽기)
        from openpyxl import load_workbook
        import io

        wb = load_workbook(io.BytesIO(excel_bytes))

        # 시트 확인
        assert "요약" in wb.sheetnames
        assert "파일 정보" in wb.sheetnames
        assert "비용 비교" in wb.sheetnames

    def test_excel_summary_sheet(self, tmp_path):
        """Excel 요약 시트 내용 확인"""
        exporter = ExcelExporter()
        calculator = PriceCalculator()

        test_file = ProcessedFile(
            file_path=tmp_path / "test.txt",
            file_type="text",
            content="Test content",
            metadata={"file_name": "test.txt", "file_size_mb": 0.001},
        )

        estimate = calculator.calculate_cost("gpt-4o-mini", input_tokens=1000)
        excel_bytes = exporter.export_workbook([test_file], [estimate], output_ratio=0.3)

        from openpyxl import load_workbook
        import io

        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["요약"]

        # 타이틀 확인
        assert ws["A1"].value == "LLM API Cost Calculator"

    def test_excel_files_sheet(self, tmp_path):
        """Excel 파일 정보 시트 내용 확인"""
        exporter = ExcelExporter()

        test_file = ProcessedFile(
            file_path=tmp_path / "test.txt",
            file_type="text",
            content="Test",
            metadata={"file_name": "test.txt", "file_size_mb": 0.001},
        )

        # 빈 estimates로 워크북 생성
        excel_bytes = exporter.export_workbook([test_file], [], output_ratio=0.3)

        from openpyxl import load_workbook
        import io

        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["파일 정보"]

        # 헤더 확인
        assert ws["A1"].value == "파일명"
        assert ws["B1"].value == "파일 타입"
        assert ws["C1"].value == "파일 크기 (MB)"

        # 데이터 확인
        assert ws["A2"].value == "test.txt"
        assert ws["B2"].value == "text"


class TestJSONExporter:
    """JSONExporter에 대한 테스트"""

    def test_export_json(self, tmp_path):
        """JSON 내보내기 테스트"""
        exporter = JSONExporter()
        calculator = PriceCalculator()

        # 테스트 파일
        test_file = ProcessedFile(
            file_path=tmp_path / "test.txt",
            file_type="text",
            content="Test content for JSON export",
            metadata={"file_name": "test.txt", "file_size_mb": 0.001},
        )

        # 비용 계산
        estimates = calculator.compare_models(
            ["gpt-4o-mini", "claude-3-haiku"], input_tokens=1000, output_tokens=500
        )

        # JSON 내보내기
        json_output = exporter.export_json([test_file], estimates, output_ratio=0.5)

        # JSON 문자열 확인
        assert isinstance(json_output, str)
        assert len(json_output) > 0

        # JSON 파싱 가능 확인
        import json

        data = json.loads(json_output)

        # 구조 확인
        assert "metadata" in data
        assert "files" in data
        assert "tokens" in data
        assert "costs" in data

    def test_json_structure(self, tmp_path):
        """JSON 데이터 구조 확인"""
        exporter = JSONExporter()
        calculator = PriceCalculator()

        test_file = ProcessedFile(
            file_path=tmp_path / "test.txt",
            file_type="text",
            content="Test",
            metadata={"file_name": "test.txt", "file_size_mb": 0.001},
        )

        estimate = calculator.calculate_cost("gpt-4o-mini", input_tokens=1000, output_tokens=500)
        json_output = exporter.export_json([test_file], [estimate], output_ratio=0.5)

        import json

        data = json.loads(json_output)

        # 메타데이터 확인
        assert data["metadata"]["tool"] == "LLM API Cost Calculator"
        assert "generated_at" in data["metadata"]

        # 파일 정보 확인
        assert data["files"]["count"] == 1
        assert len(data["files"]["items"]) == 1
        assert data["files"]["items"][0]["file_name"] == "test.txt"

        # 토큰 정보 확인
        assert "total_input_tokens" in data["tokens"]
        assert "estimated_output_tokens" in data["tokens"]
        assert data["tokens"]["output_ratio"] == 0.5

        # 비용 정보 확인
        assert data["costs"]["models_compared"] == 1
        assert "cheapest_model" in data["costs"]
        assert len(data["costs"]["all_models"]) == 1

    def test_json_with_image_file(self, tmp_path):
        """이미지 파일 JSON 내보내기 테스트"""
        exporter = JSONExporter()

        # 이미지 파일 (메타데이터 포함)
        test_file = ProcessedFile(
            file_path=tmp_path / "image.png",
            file_type="image",
            content="",
            metadata={
                "file_name": "image.png",
                "file_size_mb": 1.5,
                "width": 1024,
                "height": 768,
            },
        )

        json_output = exporter.export_json([test_file], [], output_ratio=0.3)

        import json

        data = json.loads(json_output)

        # 이미지 토큰 정보 포함 확인
        file_item = data["files"]["items"][0]
        assert file_item["file_type"] == "image"
        # image_tokens 필드가 있어야 함
        assert "image_tokens" in file_item

    def test_json_to_bytes(self):
        """JSON 바이트 변환 테스트"""
        exporter = JSONExporter()

        json_string = '{"test": "data", "korean": "한글"}'
        json_bytes = exporter.to_bytes(json_string)

        # bytes 타입 확인
        assert isinstance(json_bytes, bytes)

        # 디코딩 가능 확인
        decoded = json_bytes.decode("utf-8")
        assert decoded == json_string
        assert "한글" in decoded
